# excel.py
import os
import openpyxl

def save_credentials(email, username, password):
    filepath = "user_credentials.xlsx"
    if not os.path.exists(filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["Email", "Username", "Password"]
        sheet.append(heading)
        workbook.save(filepath)
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    sheet.append([email, username, password])
    workbook.save(filepath)

def check_credentials(username, password):
    filepath = "user_credentials.xlsx"
    if not os.path.exists(filepath):
        return False
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=3):
        if row[0].value == username and row[1].value == password:
            return True
    return False
